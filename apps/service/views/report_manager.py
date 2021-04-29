# -*- coding: utf-8 -*
from apps.service import models as mysql_db
from TestPlatform import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.http import FileResponse
from django.db.models import Avg
import json
import os
import subprocess


# 识别率测试报告
def report_manager_asr_model(**kwargs):

    # 执行测试过程

    step = kwargs["step"]

    if step == "0":

        wer, lab, rec, final_rate_list = asr_rate(step, kwargs)

        # 识别结果写进详情表（AsrModelTestResult）
        insert_table = mysql_db.AsrModelTestResult(
            test_id=kwargs["test_id"],
            wav_name=kwargs["wav_name"],
            asr_result=kwargs["result"],
            wav_answer=kwargs["wav_answer"],
            wav_len=kwargs["wav_len"],
            asr_time=kwargs["asr_time"],
            rt=kwargs["rt"],
            wer=wer,
            lab=lab,
            rec=rec,
            context=','.join(kwargs["context"]),
            wav_type=kwargs["wav_type"]
        )
        insert_table.save()

    # 处理测试结果
    if step == "1":
        # 更新测试报告表（report）
        final_rate_list = asr_rate(step, kwargs)
        final_rate_with_type = asr_rate_type(kwargs)
        insert_table_report = mysql_db.Report(test_id=kwargs["test_id"], test_details=final_rate_list, test_result=final_rate_with_type)
        try:
            insert_table_report.save()
            print kwargs["test_id"] + ": over"

        except Exception as e:
            print e


# 问真端到端测试报告
def report_manager_wz_c2c(**kwargs):
    step = kwargs["step"]
    if step == "0":
        expect_result = kwargs["expect_result"]
        run_result = kwargs["run_result"]
        # 是否开启asr识别
        if kwargs["asr_flag"]:
            wer = asr_rate(step, kwargs)[0]
            wz_asr_result_file_name = settings.REPORT_PATH + kwargs["test_id"] + "result"
            wz_asr_answer_file_name = settings.REPORT_PATH + kwargs["test_id"] + "answer"

            cmd = "rm -rf " + wz_asr_answer_file_name + " " + wz_asr_result_file_name
            os.system(cmd)

        else:
            wer = ""

        insert_table = mysql_db.WZC2CTestResult(
            test_id=kwargs["test_id"],
            wav_name=kwargs["wav_name"],
            asr_result=kwargs["result"],
            wav_answer=kwargs["wav_answer"],
            wer=wer,
            expect_result=expect_result,
            run_result=run_result,
            case_id=kwargs["case_id"],
            report=kwargs["error"],
            question_text=kwargs["question_text"],
            correct_answer=kwargs["correct_answer"],
            context=kwargs["context"],
            reference_word=kwargs["reference_word"],
            wav_type=kwargs["wav_type"],
            detail_reslut=kwargs["detail_reslut"]
        )

        insert_table.save()
    else:
        print 1

    if step == "1":

        success_count = mysql_db.WZC2CTestResult.objects.filter(report="1", test_id=kwargs["test_id"]).values().count()
        all_count = mysql_db.WZC2CTestResult.objects.filter(test_id=kwargs["test_id"]).values().count()
        try:
            rate = success_count, "/", all_count, "：", format(float(success_count)/float(all_count), '.2f')
        except:
            rate = 1
        
        insert_table_report = mysql_db.Report(test_id=kwargs["test_id"], test_details=rate, test_result=rate)
        try:
            insert_table_report.save()
            print kwargs["test_id"] + ": over"

        except Exception as e:
            print e


# 根据测试任务ID 查询测试报告
@csrf_exempt
def select_report_detail(request):
    if request.method == "POST":
        post = request.POST
        message_type = post["msgType"]
        test_id = post["testId"]
        if message_type == "AsrModelRate":
            wav_type = post["wavType"]
        else:
            wav_type = u"全部"

        limit = int(post["limit"])
        page = int(post["page"])

        if wav_type == u"全部":

            if message_type == "AsrModelRate":

                select_table_detail = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id).values()

            elif message_type == "WZC2C":
                select_table_detail = mysql_db.WZC2CTestResult.objects.filter(test_id=test_id).values()

            else:
                select_table_detail = []

        else:
            select_table_detail = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id, wav_type=wav_type).values()

        count = select_table_detail.count()

        if page != 1:
            page = ((page - 1) * limit) + 1
            limit = (limit + page) - 1

        result = list(select_table_detail[page-1:limit])

        response = {"code": 0, "count": count, "data": result}

        return JsonResponse(response)


@csrf_exempt
def select_report_main(request):
    if request.method == "POST":
        post = request.POST
        message_type = post["msgType"]
        test_id = post["testId"]
        if message_type == "AsrModelRate":
            wav_type = post["wavType"]
        else:
            wav_type = u"全部"

        if message_type == "AsrModelRate":

            report_dict = {}
            if wav_type == u"全部":
                select_main_report = mysql_db.Report.objects.filter(test_id=test_id).values()
                select_main_rt_avg = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id).values_list('test_id').annotate(avg_rt=Avg("rt")).values_list('avg_rt')

                if len(select_main_rt_avg) == 0:
                    main_rt = ""
                else:
                    main_rt = format(float(select_main_rt_avg[0][0]), '.4f')

                report_dict["rate"] = {}

                try:
                    main_wer = "WER: " + list(select_main_report)[0]["test_details"].split("Overall -> ")[-1].split("N: ")[0]
                except Exception as e:
                    print e
                    main_wer = " "
                    select_main_report = [{"test_details": ""}]

                report_dict["rate"][u"全部"] = {"rate": list(select_main_report)[0]["test_details"], "rt": main_rt, "wer": main_wer}

                try:
                    wav_type_list = eval(mysql_db.Report.objects.get(test_id=test_id).test_result).keys()
                except Exception as e:
                    print e
                    wav_type_list = []

                for i in range(len(wav_type_list)):
                    wav_type = wav_type_list[i]
                    select_type_report = eval(mysql_db.Report.objects.get(test_id=test_id).test_result)[wav_type]
                    select_type_rt_avg = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id, wav_type=wav_type).values_list('test_id').annotate(avg_rt=Avg("rt")).values_list('avg_rt')

                    type_rt = format(float(select_type_rt_avg[0][0]), '.4f')
                    type_wer = "WER: " + select_type_report.split("Overall -> ")[-1].split("N: ")[0]
                    report_dict["rate"][wav_type_list[i]] = {"rate": select_type_report, "rt": type_rt, "wer": type_wer}

                # print json.dumps(report_dict)

            else:
                report_dict["rate"] = {}

                select_type_report = eval(mysql_db.Report.objects.get(test_id=test_id).test_result)[wav_type]
                select_type_rt_avg = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id, wav_type=wav_type).values_list('test_id').annotate(avg_rt=Avg("rt")).values_list('avg_rt')
                type_rt = format(float(select_type_rt_avg[0][0]), '.4f')
                type_wer = "WER: " + select_type_report.split("Overall -> ")[-1].split("N: ")[0]

                report_dict["rate"][wav_type] = {"rate": select_type_report, "rt": type_rt, "wer": type_wer}

            try:
                response = {"code": 0, "data": json.dumps(report_dict)}
            except Exception as e:
                print e
                response = {"code": 0, "data": ""}
            return JsonResponse(response)

        elif message_type == "WZC2C":

            right_result = int(mysql_db.WZC2CTestResult.objects.filter(test_id=test_id, report="1").values().count())
            error_result = int(mysql_db.WZC2CTestResult.objects.filter(test_id=test_id, report="0").values().count())
            try:
                total_rate = format(float(right_result)/float(right_result+error_result), '.2f')
            except:
                total_rate = 1

            response = {
                "code": 0,
                "rate": total_rate,
                "right_result": right_result,
                "all_result": (right_result + error_result),
                "error_result": error_result
            }

            return JsonResponse(response)


# asr 识别率统计
def asr_rate(step, param):

    # 单独每条答案和识别结果文件（用完即删）
    result_file_name_tmp = settings.REPORT_PATH + param["test_id"] + "result_tmp"
    answer_file_name_tmp = settings.REPORT_PATH + param["test_id"] + "answer_tmp"

    # 完整答案和识别结果文件
    result_file_name = settings.REPORT_PATH + param["test_id"] + "result"
    answer_file_name = settings.REPORT_PATH + param["test_id"] + "answer"

    if step == "0":

        # 结果写进完整临时文件
        asr_result_file = open(result_file_name, "a+")
        asr_answer_file = open(answer_file_name, "a+")

        asr_result_file.write(param["wav_name"])
        asr_result_file.write(" ")
        try:
            asr_result_file.write(param["result"].encode("utf-8").replace("\n", ""))
        except:
            asr_result_file.write(param["result"].replace("\n", ""))
        asr_result_file.write(" ")

        asr_answer_file.write(param["wav_name"])
        asr_answer_file.write(" ")
        try:
            asr_answer_file.write(param["wav_answer"].encode("utf-8"))
        except:
            asr_answer_file.write(param["wav_answer"])
        asr_answer_file.write(" ")

        if param["line"] != "end":
            asr_result_file.write("\n")
            asr_answer_file.write("\n")

        asr_result_file.close()
        asr_answer_file.close()

        # 结果写入每条答案和识别结果文件

        asr_result_file_tmp = open(result_file_name_tmp, "w+")
        asr_answer_file_tmp = open(answer_file_name_tmp, "w+")

        asr_result_file_tmp.write(param["wav_name"])
        asr_result_file_tmp.write(" ")
        try:
            asr_result_file_tmp.write(param["result"].encode("utf-8").replace("\n", ""))
        except:
            asr_result_file_tmp.write(param["result"].replace("\n", ""))
        asr_result_file_tmp.write(" ")

        asr_answer_file_tmp.write(param["wav_name"])
        asr_answer_file_tmp.write(" ")
        try:
            asr_answer_file_tmp.write(param["wav_answer"].encode("utf-8"))
        except:
            asr_answer_file_tmp.write(param["wav_answer"])
        asr_answer_file_tmp.write(" ")

        asr_result_file_tmp.close()
        asr_answer_file_tmp.close()

        cmd = "sh reference/simple_eval.sh " + result_file_name_tmp + " " + answer_file_name_tmp + " " + param["test_id"]

        final = subprocess.check_output(cmd, shell=True)
        end_sign = "===========================================================================\n"

        name = final.split("\n")[0].split("utt: ")[1]

        if name.decode("utf-8") == param["wav_name"].decode("utf-8"):
            wer = "WER: " + final.split("\n")[1].split("WER: ")[1]
            lab = "lab: " + final.split("\n")[2].split("lab: ")[1]
            rec = "rec: " + final.split("\n")[3].split("rec: ")[1]
            final_rate_list = final[final.index(end_sign) + 1:-1]

        else:

            wer = "wer error"
            lab = "lab error"
            rec = "rec error"
            final_rate_list = "rate error"

        cmd = "rm -rf " + result_file_name_tmp + " " + answer_file_name_tmp + " " + param["test_id"]
        os.system(cmd)

        return wer, lab, rec, final_rate_list

    if step == "1":

        cmd = "sh reference/simple_eval.sh " + result_file_name + " " + answer_file_name + " " + param["test_id"]
        final = os.popen(cmd).readlines()
        end_sign = "===========================================================================\n"

        final_rate_list = final[final.index(end_sign)+1:-1]

        cmd = "rm -rf " + result_file_name + " " + answer_file_name
        os.system(cmd)

        return final_rate_list


# 分类识别率统计
def asr_rate_type(param):

    list_id = param["list_id"]
    test_id = param["test_id"]

    wav_type_list = mysql_db.WavType.objects.get(list_id=list_id).wav_type.split(",")
    result_dict = {}

    for i in range(len(wav_type_list)):
        result_file_name_type = settings.REPORT_PATH + test_id + wav_type_list[i] + "result"
        answer_file_name_type = settings.REPORT_PATH + test_id + wav_type_list[i] + "answer"

        result_file_type = open(result_file_name_type, "w+")
        answer_file_type = open(answer_file_name_type, "w+")

        get_asr_result = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id, wav_type=wav_type_list[i]).values()

        for j in range(len(get_asr_result)):
            wav_name = get_asr_result[j]["wav_name"]
            asr_result = get_asr_result[j]["asr_result"].encode("utf-8").replace("\n", "")
            asr_answer = get_asr_result[j]["wav_answer"].encode("utf-8").replace("\n", "")

            result_file_type.write(wav_name)
            result_file_type.write(" ")
            result_file_type.write(asr_result)
            result_file_type.write(" ")

            answer_file_type.write(wav_name)
            answer_file_type.write(" ")
            answer_file_type.write(asr_answer)
            answer_file_type.write(" ")

            if j != len(get_asr_result):
                result_file_type.write("\n")
                answer_file_type.write("\n")

        result_file_type.close()
        answer_file_type.close()

        cmd = "sh reference/simple_eval.sh " + result_file_name_type + " " + answer_file_name_type + " " + param["test_id"] + "type"
        final = subprocess.check_output(cmd, shell=True)

        end_sign = "===========================================================================\n"
        final_rate_list = final[final.index(end_sign) + 1:-1]

        result_dict[wav_type_list[i]] = final_rate_list

        cmd = "rm -rf " + result_file_name_type + " " + answer_file_name_type

        os.system(bytes(cmd.encode('utf8')))

    # update_report = mysql_db.Report.objects.get(test_id=test_id)
    # update_report.test_result = result_dict
    return result_dict


# 导出测试报告
@csrf_exempt
def export_test_report(request):
    if request.method == "POST":
        post = json.loads(request.body)
        test_id = post["testId"]
        report_type = test_id.split("-")[0]
        case_name = post["caseName"]

        if report_type == "AsrModelRate":
            export_file_list = []
            export_file_name = test_id + ".zip"
            # 全部结果
            main_report = mysql_db.Report.objects.filter(test_id=test_id).values()
            export_main_file_name = settings.REPORT_PATH + case_name + "_report"
            os.system("rm -rf " + export_main_file_name.encode("utf-8"))

            export_file = open(export_main_file_name, "a+")

            select_table = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id).values()
            print select_table
            for i in range(len(select_table)):
                wav_name_line = "utt: " + select_table[i]["wav_name"]
                wer_line = select_table[i]["wer"]
                rec_line = select_table[i]["rec"]
                lab_line = select_table[i]["lab"]

                export_file.write(wav_name_line)
                export_file.write("\n")
                export_file.write(wer_line)
                export_file.write("\n")
                export_file.write(rec_line.encode("utf-8"))
                export_file.write("\n")
                export_file.write(lab_line.encode("utf-8"))
                export_file.write("\n")

            export_file.write("===========================================================================\n")

            main_report_list = eval(main_report[0]["test_details"])

            for line in range(len(main_report_list)):
                export_file.write(main_report_list[line])

            export_file.write("===========================================================================")
            export_file.close()
            export_file_list.append(export_main_file_name)

            # 分类导出
            select_wav_type = eval(mysql_db.Report.objects.get(test_id=test_id).test_result)

            if len(select_wav_type) > 1:
                wav_type_list = select_wav_type.keys()

                for i in range(len(wav_type_list)):
                    wav_type = wav_type_list[i]
                    type_report = mysql_db.AsrModelTestResult.objects.filter(test_id=test_id, wav_type=wav_type).values()
                    export_type_file_name = settings.REPORT_PATH + case_name + wav_type + "_report"
                    os.system("rm -rf " + export_type_file_name.encode("utf-8"))

                    export_type_file = open(export_type_file_name, "a+")

                    for j in range(len(type_report)):
                        wav_name_line = "utt: " + type_report[j]["wav_name"]
                        wer_line = type_report[j]["wer"]
                        rec_line = type_report[j]["rec"]
                        lab_line = type_report[j]["lab"]

                        export_type_file.write(wav_name_line)
                        export_type_file.write("\n")
                        export_type_file.write(wer_line)
                        export_type_file.write("\n")
                        export_type_file.write(rec_line.encode("utf-8"))
                        export_type_file.write("\n")
                        export_type_file.write(lab_line.encode("utf-8"))
                        export_type_file.write("\n")

                    # export_type_file.write("===========================================================================\n")

                    type_main_report_list = select_wav_type[wav_type]

                    for line in range(len(type_main_report_list)):
                        export_type_file.write(type_main_report_list[line])

                    # export_type_file.write("===========================================================================")
                    export_type_file.close()

                    export_file_list.append(export_type_file_name)

            # 文件打包

            import zipfile

            zip_file = zipfile.ZipFile(export_file_name, 'w')

            for i in range(len(export_file_list)):
                zip_file.write(export_file_list[i], export_file_list[i].split("/")[-1])
            zip_file.close()
            # cmd = "zip -jr " + export_file_name
            # for i in range(len(export_file_list)):
            #     cmd = cmd + " " + export_file_list[i]
            # os.system(bytes(cmd.encode('utf8')))

            # print len(select_table)
            file_name = case_name
            response = FileResponse(open(export_file_name, 'rb'))

        elif report_type == "WZC2C":

            from openpyxl import Workbook
            export_file = Workbook()
            table_sheet = export_file.active

            table_sheet.append([
                "用例id",
                "音频文件名称",
                "ASR转写结果",
                "ASR标注结果",
                "ASR准确率",
                "问题",
                "系统判断结果",
                "预期判断结果",
                "预期答案",
                "测试结果",
                "详细结果",
                "reference_word",
                "context"
            ])

            select_table_detail = mysql_db.WZC2CTestResult.objects.filter(test_id=test_id).values()

            for i in range(len(select_table_detail)):
                case_id = select_table_detail[i]["case_id"]
                wav_name = select_table_detail[i]["wav_name"]
                asr_result = select_table_detail[i]["asr_result"]
                wav_answer = select_table_detail[i]["wav_answer"]
                wer = select_table_detail[i]["wer"]
                question_test = select_table_detail[i]["question_text"]
                run_result = select_table_detail[i]["run_result"]
                expect_result = select_table_detail[i]["expect_result"]
                correct_answer = select_table_detail[i]["correct_answer"]
                reference_word = select_table_detail[i]["reference_word"]
                context = select_table_detail[i]["context"]
                detail_report = select_table_detail[i]["detail_reslut"]
                if select_table_detail[i]["report"] == "1":
                    report = "成功"
                else:
                    report = "失败"

                table_sheet.append([
                    case_id,
                    wav_name,
                    asr_result,
                    wav_answer,
                    wer,
                    question_test,
                    run_result,
                    expect_result,
                    correct_answer,
                    report,
                    detail_report,
                    reference_word,
                    context
                ])
            export_file.save(settings.REPORT_PATH + case_name + ".xlsx")
            file_name = case_name + ".xlsx"
            response = FileResponse(open(settings.REPORT_PATH + case_name + ".xlsx", 'rb'))

        else:
            file_name = "none"
            response = {}

        response['Content-Type'] = 'application/octet-stream;charset=utf-8;'
        # response['Content-Disposition'] = 'attachment;filename={0}'.format(file_name.encode("utf-8"))
        response['Content-Disposition'] = "attachment;filename*=" + file_name.encode("utf-8")
        # response['Content-Length'] = "9999"
        # response['Content-Encoding'] = "UTF-8"
        # export_file.close()
        # print response
        return response

@csrf_exempt
def export_testcase(request):
    if request.method == "POST":
        post = json.loads(request.body)
        case_id = post["caseId"]
        caselist_name = post["caseName"]

        from openpyxl import Workbook
        export_file = Workbook()
        table_sheet = export_file.active

        table_sheet.append([
            "case_id",
            "case_name",
            "case_param",
            "expected",
            "method",
            "class",
            "status",
            "wav_name",
            "process",
            "reserve",
            "single_choice",
            "single_choice_question",
            "single_choice_answer",
            "question_method_id"
        ])

        select_table_detail = mysql_db.TestCase.objects.filter(file_id=case_id).values()

        for i in range(len(select_table_detail)):
            case_id = select_table_detail[i]["case_id"]
            case_name = select_table_detail[i]["case_name"]
            case_param = select_table_detail[i]["case_param"]
            expected = select_table_detail[i]["expected"]
            method = select_table_detail[i]["method"]
            testcase_class = select_table_detail[i]["test_class"]
            status = select_table_detail[i]["status"]
            wav_name = select_table_detail[i]["wav_name"]
            process = select_table_detail[i]["process"]
            reserve = select_table_detail[i]["reserve"]
            single_choice = select_table_detail[i]["single_choice"]
            single_choice_question = select_table_detail[i]["single_choice_question"]
            single_choice_answer = select_table_detail[i]["single_choice_answer"]
            question_method_id = select_table_detail[i]["question_method_id"]

            table_sheet.append([
                case_id,
                case_name,
                case_param,
                expected,
                method,
                testcase_class,
                status,
                wav_name,
                process,
                reserve,
                single_choice,
                single_choice_question,
                single_choice_answer,
                question_method_id
            ])
        export_file.save(settings.REPORT_PATH + caselist_name + ".xlsx")
        file_name = caselist_name + ".xlsx"
        response = FileResponse(open(settings.REPORT_PATH + caselist_name + ".xlsx", 'rb'))

        response['Content-Type'] = 'application/octet-stream;charset=utf-8;'
        response['Content-Disposition'] = "attachment;filename*=" + file_name.encode("utf-8")
        return response

@csrf_exempt
def export_wavlist(request):
    if request.method == "POST":
        post = json.loads(request.body)
        wav_id = post["wavId"]
        wavlist_name = post["wavName"]

        from openpyxl import Workbook
        export_file = Workbook()
        table_sheet = export_file.active

        table_sheet.append([
            "音频名称",
            "标注答案",
            "分类",
            "context",
            "reference",
            "音频地址"
        ])

        select_table_detail = mysql_db.AsrWavList.objects.filter(list_id=wav_id).values()

        for i in range(len(select_table_detail)):
            wav_name = select_table_detail[i]["wav_name"]
            wav_answer = select_table_detail[i]["wav_answer"]
            wav_type = select_table_detail[i]["wav_type"]
            wav_context = select_table_detail[i]["wav_context"]
            reference = select_table_detail[i]["reference"]
            wav_path = select_table_detail[i]["wav_path"]
            table_sheet.append([
                wav_name,
                wav_answer,
                wav_type,
                wav_context,
                reference,
                wav_path
            ])
        export_file.save(settings.REPORT_PATH + wavlist_name + ".xlsx")
        file_name = wavlist_name + ".xlsx"
        response = FileResponse(open(settings.REPORT_PATH + wavlist_name + ".xlsx", 'rb'))

        response['Content-Type'] = 'application/octet-stream;charset=utf-8;'
        response['Content-Disposition'] = "attachment;filename*=" + file_name.encode("utf-8")
        return response